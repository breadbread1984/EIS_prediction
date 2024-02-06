#!/usr/bin/python3

from absl import flags, app
from shutil import rmtree
from os import mkdir, listdir
from os.path import join, exists, splitext
import pickle
from openpyxl import load_workbook
import numpy as np
import tensorflow as tf

FLAGS = flags.FLAGS

def add_options():
  flags.DEFINE_string('input_dir', default = None, help = 'path to dataset')
  flags.DEFINE_string('output_dir', default = 'dataset', help = 'path to output directory')

def transformer():
  voltage = load_workbook(join(FLAGS.input_dir, 'Voltage.xlsx')); voltage = voltage.active
  current = load_workbook(join(FLAGS.input_dir, 'Current.xlsx')); current = current.active
  real = load_workbook(join(FLAGS.input_dir, 'EIS_real.xlsx')); real = real.active
  imag = load_workbook(join(FLAGS.input_dir, 'EIS_imag.xlsx')); imag = imag.active
  trainset_writer = tf.io.TFRecordWriter(join(FLAGS.output_dir, 'trainset.tfrecord'))
  valset_writer = tf.io.TFRecordWriter(join(FLAGS.output_dir, 'valset.tfrecord'))
  for col in range(voltage.max_column):
    v, c = list(), list()
    for row in range(voltage.max_row):
      v.append(voltage.cell(row = row + 1, column = col + 1).value)
      c.append(current.cell(row = row + 1, column = col + 1).value)
    v = np.array(v)
    c = np.array(c)
    pulse = np.stack([v,c], axis = -1) # pulse.shape = (length, 2)
    r, i = list(), list()
    for row in range(real.max_row):
      r.append(real.cell(row = row + 1, column = col + 1).value)
      i.append(imag.cell(row = row + 1, column = col + 1).value)
    r = np.array(r)
    i = np.array(i)
    eis = np.stack([r,i], axis = -1) # eis.shape = (length, 2)
    trainsample = tf.train.Example(features = tf.train.Features(
      feature = {
        'x': tf.train.Feature(bytes_list = tf.train.BytesList(value = [tf.io.serialize_tensor(pulse).numpy()])),
        'y': tf.train.Feature(bytes_list = tf.train.BytesList(value = [tf.io.serialize_tensor(eis).numpy()]))
      }))
    is_train = np.random.multinomial(1, (9/10,1/10), size = ())[0]
    if is_train:
      trainset_writer.write(trainsample.SerializeToString())
    else:
      valset_writer.write(trainsample.SerializeToString())
  trainset_writer.close()
  valset_writer.close()

def main(unused_argv):
  if exists(FLAGS.output_dir): rmtree(FLAGS.output_dir)
  mkdir(FLAGS.output_dir)
  transformer()

if __name__ == "__main__":
  add_options()
  app.run(main)


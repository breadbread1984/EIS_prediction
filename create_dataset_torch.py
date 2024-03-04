#!/usr/bin/python3

from absl import flags, app
from shutil import rmtree
from os import mkdir, listdir
from os.path import join, exists, splitext
import pickle
from openpyxl import load_workbook
import numpy as np
from torch.utils.data import Dataset, DataLoader
from uuid import uuid1

FLAGS = flags.FLAGS

def add_options():
  flags.DEFINE_string('input_dir', default = None, help = 'path to dataset')
  flags.DEFINE_string('output_dir', default = 'dataset', help = 'path to output directory')

def main(unused_argv):
  if exists(FLAGS.output_dir): rmtree(FLAGS.output_dir)
  mkdir(FLAGS.output_dir)
  mkdir(join(FLAGS.output_dir, 'train'))
  mkdir(join(FLAGS.output_dir, 'val'))
  voltage = load_workbook(join(FLAGS.input_dir, 'Voltage.xlsx')); voltage = voltage.active
  current = load_workbook(join(FLAGS.input_dir, 'Current.xlsx')); current = current.active
  real = load_workbook(join(FLAGS.input_dir, 'EIS_real.xlsx')); real = real.active
  imag = load_workbook(join(FLAGS.input_dir, 'EIS_imag.xlsx')); imag = imag.active
  for col in range(voltage.max_column):
    v, c = list(), list()
    for row in range(voltage.max_row):
      v.append(voltage.cell(row = row + 1, column = col + 1).value)
      c.append(current.cell(row = row + 1, column = col + 1).value)
    v = np.array(v)
    c = np.array(c)
    pulse = np.stack([v,c], axis = -1) # pulse.shape = (length, 2)
    r, i = list(), list()
    for row in range(real.max_row):
      r.append(real.cell(row = row + 1, column = col + 1).value)
      i.append(imag.cell(row = row + 1, column = col + 1).value)
    r = np.array(r)
    i = np.array(i)
    eis = np.stack([r,i], axis = -1) # eis.shape = (length, 2)
    is_train = np.random.multinomial(1, (9/10,1/10), size = ())[0]
    dir_path = 'train' if is_train else 'val'
    np.savez(join(FLAGS.output_dir, dir_path, str(uuid1()) + '.npz'), x = pulse, y = eis)

class EISDataset(Dataset):
  def __init__(self, dataset_dir):
    super(EISDataset, self).__init__()
    self.file_list = list()
    for f in listdir(dataset_dir):
      stem, ext = splitext(f)
      if ext != '.npz': continue
      self.file_list.append(join(dataset_dir, f))
  def __len__(self):
    return len(self.file_list)
  def __getitem__(self, idx):
    data = np.load(self.file_list[idx])
    x, y = data['x'].astype(np.float32), data['y'].astype(np.float32)
    y = np.stack([np.log(y[:,0]),np.exp(y[:,1])], axis = -1)
    return x, y

if __name__ == "__main__":
  add_options()
  app.run(main)

